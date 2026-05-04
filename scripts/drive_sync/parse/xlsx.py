"""xlsx → list[FacultyGroup].

Sheet shape (single sheet, header in row 1):
    program | degree | faculty | department | credits | duration | language | source

Header names are case-insensitive and trimmed. `program`, `degree`, `faculty`
are required. Rows are grouped by `faculty`; the heading is parsed for the
`<Name> ([<Abbr>](<url>))` form to extract abbreviation and link.

Design: spec/001-add-google-drive-backend-data/design.md §3.4.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from pydantic import HttpUrl, ValidationError

from drive_sync.models import FacultyGroup, MajorRow
from drive_sync.report import ParseReport


REQUIRED_COLS = ("program", "degree", "faculty")
ALL_COLS = ("program", "degree", "faculty", "department", "credits", "duration", "language", "source")


_FACULTY_HEADING_RE = re.compile(
    r"^(?P<heading>.+?)\s*\(\s*\[(?P<abbr>[^\]]+)\]\((?P<url>https?://[^\s)]+)\)\s*\)\s*$"
)


@dataclass
class ParseXlsxOptions:
    file_label: str
    web_view_link: str | None = None


def parse_xlsx(
    file_path: str,
    options: ParseXlsxOptions,
    report: ParseReport,
) -> list[FacultyGroup] | None:
    """Parse an xlsx file into ordered FacultyGroups.

    Returns None on fatal errors (recorded in the report).
    """
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as err:  # noqa: BLE001
        report.error(
            options.file_label,
            f"cannot open xlsx: {err}",
            web_view_link=options.web_view_link,
        )
        return None

    if not wb.worksheets:
        report.error(
            options.file_label,
            "xlsx contains no worksheets",
            web_view_link=options.web_view_link,
        )
        return None

    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows()
    try:
        header_row = next(rows_iter)
    except StopIteration:
        report.error(
            options.file_label,
            "xlsx is empty (no header row)",
            web_view_link=options.web_view_link,
        )
        return None

    header_by_col: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        raw = _cell_string(cell).strip().lower()
        if raw in ALL_COLS:
            header_by_col[idx] = raw

    for required in REQUIRED_COLS:
        if required not in header_by_col.values():
            report.error(
                options.file_label,
                f'xlsx is missing required column "{required}"',
                web_view_link=options.web_view_link,
            )
            return None

    # Iterate data rows. row_num here is the 1-indexed sheet row (matches Excel).
    raw_rows: list[tuple[int, dict[str, str]]] = []
    for row_idx, row in enumerate(rows_iter, start=2):
        raw: dict[str, str] = {}
        has_any = False
        for col_idx, name in header_by_col.items():
            v = _cell_string(row[col_idx]).strip() if col_idx < len(row) else ""
            if v:
                has_any = True
            raw[name] = v
        if not has_any:
            continue
        raw_rows.append((row_idx, raw))

    if not raw_rows:
        # An empty sheet with just a header isn't necessarily an error — let
        # the caller decide. For now we return an empty list.
        return []

    # Validate + group by faculty, preserving first-seen order.
    groups: dict[str, list[MajorRow]] = {}
    group_order: list[str] = []
    had_errors = False

    for row_num, raw in raw_rows:
        try:
            row_model = _build_major_row(raw)
        except ValidationError as err:
            had_errors = True
            issues = "; ".join(
                f"{'.'.join(str(p) for p in e['loc']) or '(row)'}: {e['msg']}"
                for e in err.errors()
            )
            report.error(
                options.file_label,
                f"row {row_num} invalid — {issues}",
                web_view_link=options.web_view_link,
                where=f"row {row_num}",
            )
            continue
        except ValueError as err:
            had_errors = True
            report.error(
                options.file_label,
                f"row {row_num} invalid — {err}",
                web_view_link=options.web_view_link,
                where=f"row {row_num}",
            )
            continue

        faculty_key = raw["faculty"]
        if faculty_key not in groups:
            groups[faculty_key] = []
            group_order.append(faculty_key)
        groups[faculty_key].append(row_model)

    if had_errors:
        return None

    return [
        _build_faculty_group(facu, groups[facu]) for facu in group_order
    ]


def parse_faculty_heading(text: str) -> tuple[str, str | None, str | None]:
    """Split `<Name> ([<Abbr>](<url>))` into (heading, abbr, url).

    Falls back to the whole string as heading if the pattern doesn't match.
    """
    m = _FACULTY_HEADING_RE.match(text)
    if not m:
        return text, None, None
    return m.group("heading").strip(), m.group("abbr"), m.group("url")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _build_major_row(raw: dict[str, str]) -> MajorRow:
    """Construct a MajorRow from the raw cell strings.

    Translates the xlsx column names to the IR field names:
    - `duration` (xlsx) → `years` (IR)
    - empty strings → None for optional fields
    """
    if not raw.get("program"):
        raise ValueError("program is required")
    if not raw.get("degree"):
        raise ValueError("degree is required")

    def _opt(key: str) -> str | None:
        v = raw.get(key, "")
        return v if v else None

    def _opt_int(key: str) -> int | None:
        v = raw.get(key, "")
        if not v:
            return None
        try:
            return int(v)
        except ValueError as err:
            raise ValueError(f"{key} must be an integer, got {v!r}") from err

    return MajorRow(
        program=raw["program"],
        degree=raw["degree"],
        department=_opt("department"),
        credits=_opt_int("credits"),
        years=_opt_int("duration"),
        language=_opt("language"),
        source=_opt("source"),  # pydantic HttpUrl will validate
    )


def _build_faculty_group(faculty_text: str, rows: list[MajorRow]) -> FacultyGroup:
    heading, abbr, url = parse_faculty_heading(faculty_text)
    return FacultyGroup(heading=heading, abbr=abbr, url=url, rows=rows)


def _cell_string(cell: Cell | None) -> str:
    """Coerce a cell value to a string, handling None / numeric / hyperlink cells."""
    if cell is None:
        return ""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    if isinstance(v, (int, float)):
        # Print integers without trailing .0 for cleanliness.
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    if isinstance(v, bool):
        return "true" if v else "false"
    return str(v)


# Suppress unused-import warning.
_ = (Iterable, HttpUrl)
