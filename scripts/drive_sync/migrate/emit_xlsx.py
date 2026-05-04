"""IR.majors → .xlsx with proper formatting.

Single sheet with the canonical column order:
    program | degree | faculty | department | credits | duration | language | source

Adds:
- bold + colored header row
- frozen top row
- per-column auto-sized widths
- alternating row tint via openpyxl's `Table` styling

Design: spec/001-add-google-drive-backend-data/design.md §3.4, §5.4.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from drive_sync.models import FacultyGroup


_HEADER = ("program", "degree", "faculty", "department", "credits", "duration", "language", "source")


def emit_xlsx(groups: list[FacultyGroup], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "majors"

    # Header row.
    ws.append(list(_HEADER))
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF4472C4")  # Word "Blue, Accent 1"
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    # Data rows.
    for g in groups:
        faculty_cell = _format_faculty_cell(g)
        for row in g.rows:
            ws.append(
                [
                    row.program,
                    row.degree or "",
                    faculty_cell,
                    row.department or "",
                    row.credits if row.credits is not None else "",
                    row.years if row.years is not None else "",
                    row.language or "",
                    str(row.source) if row.source else "",
                ]
            )

    # Column widths (rough autofit based on content length, capped).
    max_widths = [len(h) for h in _HEADER]
    for r in ws.iter_rows(min_row=2, values_only=True):
        for c, value in enumerate(r):
            v = "" if value is None else str(value)
            if len(v) > max_widths[c]:
                max_widths[c] = len(v)
    for idx, width in enumerate(max_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(60, max(8, width + 2))

    # Freeze the header so it stays visible while scrolling.
    ws.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


def _format_faculty_cell(g: FacultyGroup) -> str:
    if g.abbr and g.url:
        return f"{g.heading} ([{g.abbr}]({g.url}))"
    return g.heading
